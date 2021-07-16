import openpyxl
import json
from os import path
from datetime import date

file_name = '商家資料填寫表格.xlsx'


def parse_food_items(sheet):
    item_map = {}
    reverse_idx = {}

    kind_ary = []
    item_count_map = {}

    for row in sheet.iter_rows(min_row=2, max_col=4):
        food_name = row[0].value
        kind = row[1].value
        price = row[2].value
        memo = row[3].value
        if food_name is None or kind is None or price is None:
            print(f'有資料空缺，此筆未寫入。{kind}, {food_name}, {price}')
            continue

        # 產生 item_key，並初始化類別
        try:
            kind_idx = kind_ary.index(kind)
        except ValueError:
            kind_idx = len(kind_ary)
            kind_ary.append(kind)

        food_count = item_count_map.get(kind_idx, 0)
        item_key = f'{kind_idx}_{food_count}'

        # 存入map
        kind = item_map.get(str(kind_idx), {'kindname': kind})
        items = kind.get('items', {})
        items[item_key] = {'foodname': food_name, 'price': price}
        if memo:
            items[item_key]['memo'] = memo

        kind['items'] = items
        item_map[str(kind_idx)] = kind

        # 更新 reverse index
        reverse_idx[food_name] = item_key
        # 更新 item_count_map
        item_count_map[kind_idx] = food_count + 1

    return item_map, reverse_idx

    pass


def parse_taste(sheet, item_idx_map):
    taste_map = {}
    taste_count_map = {}

    for row in sheet.iter_rows(min_row=2, max_col=3):
        taste = row[0].value
        item_name = row[1].value
        price = row[2].value

        if taste is None or item_name is None or price is None:
            print(f'有資料空缺，此筆未寫入。{taste}, {item_name}, {price}')
            continue

        if item_name in item_idx_map:
            # 取得 口味
            item_key = item_idx_map[item_name]
            taste_count = taste_count_map.get(item_name, 0)
            taste_key = f'{item_key}_{taste_count}'

            # 加入 map
            taste_map_in_item = taste_map.get(item_key, {})
            taste_map_in_item[taste_key] = {'price': price, 'tasteName': taste}
            taste_map[item_key] = taste_map_in_item

            # 更新口味數
            taste_count_map[item_name] = taste_count + 1

    return taste_map
    pass

def parse_shop(sheet):
    name = sheet['B2'].value
    phone = sheet['B4'].value
    compony_id = sheet['B3'].value
    address = sheet['B6'].value or sheet['B5'].value
    sid = sheet['D2'].value

    if sid is None or sid == "":
        raise ValueError('無店家 ID')

    return {
        'name' : name,
        'phone': phone,
        'compony_id': compony_id,
        'address': address,
        'sid': sid
    }

def parse_shop_rows(sheet):
    obj = {}

    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=3):
        key = row[0].value
        value = row[1].value or ""
        if key is None:
            print('店家資料不完整')
        else:
            obj[key] = value
        
    return obj

def parse_excel(excel_name):
    excel = openpyxl.load_workbook(excel_name)

    item_sheet = excel['品項']
    food_items, revere_index = parse_food_items(item_sheet)

    taste_sheet = excel['口味']
    tastes = parse_taste(taste_sheet, revere_index)

    shop_sheet = excel['商家資料表格']
    shop_info = parse_shop_rows(shop_sheet)

    return {'kinds': food_items, 'taste': tastes, 'shop': shop_info}

    pass

def save_shop_info(data):
    
    name = data['shop']['name'] or 'noname'
    sid = data['shop']['sid']
    print(name,sid)

    with open(f'{sid}_{name}_{date.today().isoformat()}.json', 'w', encoding='utf8') as f:
        json.dump(data, f, ensure_ascii=False)

    pass


def add_file_postfix(name):
    if name[-4:] == '.xls':
        raise Exception('不支援舊格式 excel')

    if name[-5:] != '.xlsx':
        name = name + '.xlsx'

    return name
    pass


def ask_file_name(name: str):
    print('請輸入 excel 檔名，或是 enter 使用預設值:')
    user_input = input(f'({name}): ') or name

    with_postfix = add_file_postfix(user_input)

    if not path.exists(with_postfix):
        raise ValueError('尋無檔案')

    return with_postfix
    pass


if __name__ == "__main__":
    try:
        fn = ask_file_name(file_name)
        shop_info = parse_excel(fn)
        save_shop_info(shop_info)
        input('轉檔完成')
    except Exception as e:
        input(e) 
        pass